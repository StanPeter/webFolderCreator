import os
import zipfile

import openpyxl as op


# keep the database.xlsx file in the same directory as the web_creator.py file
file = 'database.xlsx'
wb = op.load_workbook(file)
sheet = wb['Sheet1']
max_row = sheet.max_row
# the file will be created in the same folder as web_creator.py
path = os.getcwd() + '/'
# path = os.chdir('set/your/own/path/')     #change for creating an own path


def file_structure():
    print('step 1')
    for i in range(2, max_row):
        try:
            os.makedirs(str(sheet.cell(column=1, row=i).value))
        except OSError:
            pass

    print('step 2')
    for i in range(2, max_row):
        os.chdir(path + str(sheet.cell(column=1, row=i).value))
        try:
            os.makedirs(str(sheet.cell(column=2, row=i).value))
        except OSError:
            pass

    print('step 3')
    for i in range(2, max_row):
        os.chdir(path + str(sheet.cell(column=1, row=i).value) +
                 '/' + str(sheet.cell(column=2, row=i).value))
        try:
            os.makedirs(str(sheet.cell(column=4, row=i).value))
        except OSError:
            pass

    print('step 4')
    for i in range(2, max_row):
        os.chdir(path + str(sheet.cell(column=1, row=i).value) + '/' + str(sheet.cell(
            column=2, row=i).value) + '/' + str(sheet.cell(column=4, row=i).value))
        try:
            os.makedirs(str(sheet.cell(column=6, row=i).value))
        except OSError:
            pass

    print('step 5')
    for i in range(2, max_row):
        os.chdir(path + str(sheet.cell(column=1, row=i).value) + '/' + str(sheet.cell(column=2, row=i).value) +
                 '/' + str(sheet.cell(column=4, row=i).value) + '/' + str(sheet.cell(column=6, row=i).value))
        try:
            os.makedirs(str(sheet.cell(column=8, row=i).value))
        except OSError:
            pass

    print('step 6')
    for i in range(2, max_row):
        os.chdir(path + str(sheet.cell(column=1, row=i).value) + '/' + str(sheet.cell(column=2, row=i).value) + '/' + str(sheet.cell(
            column=4, row=i).value) + '/' + str(sheet.cell(column=6, row=i).value) + '/' + str(sheet.cell(column=8, row=i).value))
        try:
            os.makedirs(str(sheet.cell(column=10, row=i).value))
        except OSError:
            pass

    print('step 7')
    for i in range(2, max_row):
        os.chdir(path + str(sheet.cell(column=1, row=i).value) + '/' + str(sheet.cell(column=2, row=i).value) + '/' + str(sheet.cell(column=4, row=i).value) +
                 '/' + str(sheet.cell(column=6, row=i).value) + '/' + str(sheet.cell(column=8, row=i).value) + '/' + str(sheet.cell(column=10, row=i).value))
        try:
            os.makedirs(str(sheet.cell(column=11, row=i).value))
        except OSError:
            pass

        os.chdir(path + str(sheet.cell(column=1, row=i).value) + '/' + str(
            sheet.cell(column=2, row=i).value) + '/' + str(sheet.cell(column=4, row=i).value) + '/' + str(
            sheet.cell(column=6, row=i).value) + '/' + str(sheet.cell(column=8, row=i).value) + '/' + str(
            sheet.cell(column=10, row=i).value) + '/' + str(sheet.cell(column=11, row=i).value))

        try:
            open('read_me.txt', 'w+')
            zf = zipfile.ZipFile(
                (sheet.cell(column=14, row=i).value)+'.zip', mode='w')
            try:
                zf.write('README.txt')
                zf.close()
            except RuntimeError:
                pass
        except OSError:
            pass


file_structure()
